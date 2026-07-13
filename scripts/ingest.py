#!/usr/bin/env python3
"""
SAO Ingestion Engine
Processes files inside `raw/` and outputs clean Markdown under `wiki/`
Moves original files to `ingested/` and triggers graphify update.
"""

import sys
import os
import json
import shutil
import re
from datetime import datetime
from pathlib import Path
import urllib.request

CONFIG_PATH = os.path.expanduser("~/.sao/config.json")
HERMES_PORT = 20477

# Text extractors
try:
    import docx
except ImportError:
    docx = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


def load_vault_path():
    if not os.path.exists(CONFIG_PATH):
        return None
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f).get("vault_path")
    except Exception:
        return None


def extract_txt(path):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception as e:
        return f"[Error reading text file: {e}]"


def extract_docx(path):
    if not docx:
        return "[python-docx not installed. Cannot parse .docx]"
    try:
        doc = docx.Document(path)
        return "\n".join([p.text for p in doc.paragraphs])
    except Exception as e:
        return f"[Error reading docx: {e}]"


def extract_xlsx(path):
    if not openpyxl:
        return "[openpyxl not installed. Cannot parse .xlsx]"
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        out = []
        for name in wb.sheetnames[:3]:
            out.append(f"Sheet: {name}")
            ws = wb[name]
            for row in ws.iter_rows(max_row=50, values_only=True):
                if any(row):
                    out.append(", ".join([str(v) if v is not None else "" for v in row]))
        return "\n".join(out)
    except Exception as e:
        return f"[Error reading xlsx: {e}]"


def query_llm(system_prompt, user_prompt):
    """Call Hermes core (port 20477) for formatting raw documents."""
    url = f"http://localhost:{HERMES_PORT}/v1/chat/completions"
    payload = {
        "model": "fusion",
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        "temperature": 0.2
    }
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=90) as res:
            resp = json.loads(res.read().decode("utf-8"))
            return resp["choices"][0]["message"]["content"]
    except Exception as e:
        print(f"⚠️ API Call failed: {e}. Check if SAO is running ('sao start').")
        return None


def clean_filename(title):
    t = title.lower().strip()
    t = re.sub(r"[^a-z0-9_\-\s]", "", t)
    t = re.sub(r"[\s_-]+", "-", t)
    return t[:50]


def process_file(file_path, wiki_dir):
    suffix = file_path.suffix.lower()
    content = ""
    
    print(f"Reading {file_path.name} ...")
    if suffix in [".txt", ".md", ".html", ".csv", ".json", ".yaml", ".yml"]:
        content = extract_txt(file_path)
    elif suffix == ".docx":
        content = extract_docx(file_path)
    elif suffix in [".xlsx", ".xls"]:
        content = extract_xlsx(file_path)
    else:
        # Unsupported / binary. Let user know.
        print(f"⚠️ Unsupported binary file: {file_path.name}. Skip parsing, create placeholder.")
        content = f"[Unparsed binary attachment: {file_path.name}]"

    if not content.strip():
        print(f"⚪ Empty file skipped: {file_path.name}")
        return False

    system_prompt = (
        "You are the SAO Ingestion Engine. Your task is to transform raw document content into a CLEAN Obsidian Markdown file.\n"
        "Rules:\n"
        "1. Write yaml frontmatter: title, date (YYYY-MM-DD), type (concept/project/reference/etc), status (canonical/hypothesis).\n"
        "2. Add tags, specifically including 'ingested' and domain tags.\n"
        "3. Rewrite the body using neat headers. Clean up typos, formatting errors, or messy transcripts.\n"
        "4. DO NOT change technical details, code blocks, or names.\n"
        "5. Intelligently add [[wikilinks]] pointing to other potential topics in the vault.\n"
        "6. Do not include markdown outer blocks (like ```markdown), output raw markdown immediately starting with ---."
    )

    user_prompt = f"Original filename: {file_path.name}\n\nContent:\n{content}"
    
    print("Sending to Sira for translation into Clean structure...")
    formatted = query_llm(system_prompt, user_prompt)
    if not formatted:
        return False

    # Extract title from frontmatter or llm response
    title_match = re.search(r'^title:\s*"([^"]+)"', formatted, re.M)
    title = title_match.group(1) if title_match else file_path.stem
    
    clean_name = clean_filename(title)
    if not clean_name:
        clean_name = file_path.stem
        
    out_file = wiki_dir / f"{clean_name}.md"
    
    # Avoid overwriting
    counter = 1
    while out_file.exists():
        out_file = wiki_dir / f"{clean_name}-{counter}.md"
        counter += 1

    out_file.write_text(formatted, encoding="utf-8")
    print(f"✅ Ingested to wiki: {out_file.name}")
    return True


def run_ingestion():
    vpath = load_vault_path()
    if not vpath or not os.path.exists(vpath):
        print("❌ Vault path not set or invalid. Run 'sao setup vault' first.")
        sys.exit(1)

    vault = Path(vpath)
    raw_dir = vault / "raw"
    ingested_dir = vault / "ingested"
    wiki_dir = vault / "wiki"

    raw_dir.mkdir(parents=True, exist_ok=True)
    ingested_dir.mkdir(parents=True, exist_ok=True)
    wiki_dir.mkdir(parents=True, exist_ok=True)

    files = [f for f in raw_dir.iterdir() if f.is_file() and f.name != ".gitkeep" and not f.name.startswith(".")]

    if not files:
        print("⚪ No files found in raw/ to ingest.")
        return

    print(f"📦 Found {len(files)} files to ingest. Processing...")
    
    success_count = 0
    for f in files:
        try:
            success = process_file(f, wiki_dir)
            if success:
                # Move to ingested
                dest = ingested_dir / f.name
                # Avoid overwrite
                if dest.exists():
                    dest = ingested_dir / f"{f.stem}_{int(datetime.now().timestamp())}{f.suffix}"
                shutil.move(str(f), str(dest))
                success_count += 1
        except Exception as e:
            print(f"❌ Error processing {f.name}: {e}")

    print(f"\n🎉 Ingestion complete: {success_count}/{len(files)} files processed successfully.")

    # Trigger graphify update
    if success_count > 0:
        print("🔄 Updating Graphify Index...")
        python_bin = sys.executable
        graphify_dir = Path(__file__).resolve().parent.parent / "services" / "graphify"
        if (graphify_dir / ".venv").exists():
            if sys.platform == "win32":
                python_bin = str(graphify_dir / ".venv" / "Scripts" / "python.exe")
            else:
                python_bin = str(graphify_dir / ".venv" / "bin" / "python")

        import subprocess
        try:
            subprocess.run([python_bin, "-m", "graphify", "update", vpath], cwd=str(graphify_dir), check=True)
            print("✅ Graphify Index updated successfully.")
        except Exception as e:
            print(f"⚠️ Failed to update graphify index: {e}")


if __name__ == "__main__":
    run_ingestion()
